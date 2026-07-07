"""Сервис экспорта данных в Excel"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from app.db.models import User, Result, Feedback, Referral
from datetime import datetime
from pathlib import Path


class ExcelExportService:
    """Сервис экспорта данных в Excel"""
    
    @staticmethod
    async def export_users(session: AsyncSession, output_path: str = None) -> str:
        """Экспортирует пользователей в Excel"""
        if output_path is None:
            output_path = f"export_users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Заголовки
        headers = ["ID", "Telegram ID", "Username", "Phone", "Language", "Created At"]
        ws.append(headers)
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Получаем данные
        stmt = select(User)
        result = await session.execute(stmt)
        users = result.scalars().all()
        
        # Добавляем данные
        for user in users:
            ws.append([
                user.id,
                user.telegram_id,
                user.username,
                user.phone,
                user.language,
                user.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        # Форматирование ширины колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    async def export_results(session: AsyncSession, output_path: str = None) -> str:
        """Экспортирует результаты в Excel"""
        if output_path is None:
            output_path = f"export_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Заголовки
        headers = ["ID", "Telegram ID", "User ID", "Blue", "Green", "Yellow", "Red", 
                  "Primary", "Secondary", "Completed", "Created At"]
        ws.append(headers)
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Получаем данные
        stmt = select(Result)
        result = await session.execute(stmt)
        results = result.scalars().all()
        
        # Добавляем данные
        for res in results:
            ws.append([
                res.id,
                res.telegram_id,
                res.user_id,
                res.blue_score,
                res.green_score,
                res.yellow_score,
                res.red_score,
                res.primary_color,
                res.secondary_color,
                "Yes" if res.completed else "No",
                res.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        # Форматирование ширины колонок
        for i, col in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']):
            ws.column_dimensions[col].width = 15
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    async def export_feedback(session: AsyncSession, output_path: str = None) -> str:
        """Экспортирует отзывы в Excel"""
        if output_path is None:
            output_path = f"export_feedback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        
        # Заголовки
        headers = ["ID", "Telegram ID", "Result ID", "Rating", "Comment", "Agree Referral", "Created At"]
        ws.append(headers)
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Получаем данные
        stmt = select(Feedback)
        result = await session.execute(stmt)
        feedbacks = result.scalars().all()
        
        # Добавляем данные
        for fb in feedbacks:
            ws.append([
                fb.id,
                fb.telegram_id,
                fb.result_id,
                fb.rating,
                fb.comment or "",
                "Yes" if fb.agree_referral else "No",
                fb.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        # Форматирование ширины колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    async def export_referrals(session: AsyncSession, output_path: str = None) -> str:
        """Экспортирует рефералы в Excel"""
        if output_path is None:
            output_path = f"export_referrals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Referrals"
        
        # Заголовки
        headers = ["ID", "Referrer ID", "Referred ID", "Completed Test", "Created At"]
        ws.append(headers)
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Получаем данные
        stmt = select(Referral)
        result = await session.execute(stmt)
        referrals = result.scalars().all()
        
        # Добавляем данные
        for ref in referrals:
            ws.append([
                ref.id,
                ref.referrer_id,
                ref.referred_id,
                "Yes" if ref.completed_test else "No",
                ref.created_at.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        # Форматирование ширины колонок
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20
        
        wb.save(output_path)
        return output_path
